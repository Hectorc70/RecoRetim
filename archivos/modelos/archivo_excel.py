#import pdb; pdb.set_trace()

import openpyxl
from openpyxl import load_workbook
from openpyxl import workbook


class Archivo_excel():

	"""Lee el archivo excel"""

	def __init__(self, documento):

		self.ruta = documento
		

		self.abrir_documento()
		self.obtener_hojas()


	def abrir_documento(self):
		self.documento_open =  load_workbook(self.ruta)
		print("Se Abrio el documento: ",
		 	  self.ruta.split("\\")[-1])




	def obtener_hojas(self):
		"""Obtiene los nombres de las hojas
			del archivo excel dado"""

		self.hojas_nombres = {}

		for hoja_nombre, hojas in  zip(self.documento_open.sheetnames,
									   self.documento_open.worksheets):
			self.hojas_nombres[hoja_nombre] = hojas

		self.hojas_lista = [h for h in self.documento_open.worksheets]

		return(self.hojas_nombres)




	def leer_titulos(self, hoja, linea):
		"""Obtiene titulos de columnas del
			archivo que se abre"""

		self.linea_i = linea
		self.linea_ccn = self.linea_i -1
		self.titulos_columnas = {}
		self.hoja = hoja 
		#LLama al metodo que obtiene los titulos cccn del IQ
		if self.ruta.split("\\")[-1].split("_")[0] == 'IQ':
			self.leer_titulos_ccn_iq(hoja, self.linea_ccn)


		if type(self.hoja) is str:

			hoja = self.hojas_nombres[self.hoja]
			titulos = hoja[self.linea_i]

			for titulo in range(len(titulos)):
				if titulos[titulo].value is not None:
					self.titulos_columnas[titulos[titulo].value] = titulos[titulo]



		elif type(self.hoja) is int:

			hoja = self.hojas_lista[self.hoja]
			titulos = hoja[self.linea_i]


			for titulo in range(len(titulos)):
				if titulos[titulo].value is not None:
					self.titulos_columnas[titulos[titulo].value] = titulos[titulo]

		self.obtener_claves_celdas(self.hoja)



	def obtener_claves_celdas(self, hoja):

		self.claves_columnas     = {}
		self.columnas_i = {}

		if type(hoja) is str :
			columna_max = self.hojas_nombres[hoja].max_column
			fila_max 	 = self.hojas_nombres[hoja].max_row

			for (celda_text,
				 celda_nombres_clv) in self.titulos_columnas.items():

					celda          = str(celda_nombres_clv).split(".")[-1][:-2]
					self.claves_columnas[celda_text.strip(' ')]  = celda.strip('>')		#Clave de la columna 					
					self.columnas_i [celda_text.strip(' ')] = self.linea_i			#fila inicial de lectura

		elif type(hoja) == int and type(hoja) != float:
			columna_max = self.hojas_lista[hoja].max_column
			fila_max 	 = self.hojas_lista[hoja].max_row

			for (celda_text, celda_nombres_clv) in self.titulos_columnas.items():

					celda          = str(celda_nombres_clv).split(".")[-1][:-2]
					self.claves_columnas[celda_text.strip(' ')]  = celda.strip('>')		#Clave de la columna
					self.columnas_i [celda_text.strip(' ')] = self.linea_i			#Fila donde se empieza a leer







	def leer_titulos_ccn_iq(self, hoja_lectura, linea):
		"""Obtiene los titulos del IQ NUMEROS DE CONCEPTOS
		   de Nómina"""
		self.titulos_ccn = {}

		

		if type(hoja_lectura) is str:
			hoja = self.hojas_nombres[hoja_lectura]
			titulos = hoja[linea]

			for titulo in range(len(titulos)):
				if titulos[titulo].value is not None:
					self.titulos_ccn[titulos[titulo].value] = titulos[titulo]



		elif type(hoja_lectura) is int:

			hoja = self.hojas_lista[hoja_lectura]
			titulos = hoja[linea]

			for titulo in range(len(titulos)):
				if titulos[titulo].value is not None:
					self.titulos_ccn[titulos[titulo].value] = titulos[titulo]


		self.obtener_claves_celdas_ccn(hoja_lectura)


	def obtener_claves_celdas_ccn(self, hoja):

		self.columnas_ccn     = {}
		self.columnas_i_ccn = {}



		if type(hoja) is str :
			columna_max = self.hojas_nombres[hoja].max_column
			fila_max 	 = self.hojas_nombres[hoja].max_row

			for (celda_text,
				 celda_nombres_clv) in self.titulos_ccn.items():

				celd = str(celda_text)

				celda= str(celda_nombres_clv).split(".")[-1][:-2]
				self.columnas_ccn[celd.strip(' ')]  = celda.strip('>')   #Claves de las columnas				
				self.columnas_i_ccn [celd.strip(' ')] = self.linea_i	#fila inicial de lectura

		elif type(hoja) == int and type(hoja) != float:
			columna_max = self.hojas_lista[hoja].max_column
			fila_max 	 = self.hojas_lista[hoja].max_row

			for (celda_text, celda_nombres_clv) in self.titulos_ccn.items():

				celd = str(celda_text)
				celda          = str(celda_nombres_clv).split(".")[-1][:-2]
				self.columnas_ccn[celd.strip(' ')]  = celda.strip('>')		#Claves de las columnas				
				self.columnas_i_ccn [celd.strip(' ')] = self.linea_i		#fila inicial de lectura
		

