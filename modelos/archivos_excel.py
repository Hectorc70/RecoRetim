import os.path as path
from tkinter.filedialog import askopenfilename

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

def comprobar(ruta):

	if path.exists(ruta):
		ArchivoExcelAbrir(ruta)

	else:
		ArchivoExcelNuevo(ruta)

class ArchivoExcelEscribir:

	def __init__(self, ruta, titulo_hoja_default = "Hoja 1"):

		self.abrir_documento(ruta)
		self.wb.active       = 0
		self.wb.active.title = str(titulo_hoja_default)  



	def crear_archivo_xlsx(self):
		"""Crea un archivo nuevo, excel 'xlsx'"""

		self.wb = Workbook()

	
	def abrir_documento(self, ruta):
		self.wb =  load_workbook(ruta)
		print("Se Abrio el documento: ",
		 	  self.ruta.split("\\")[-1])


	def crear_hoja_nueva(self, nombre_hoja_nueva):
		"""Crea una hoja nueva en el archivo 
		   excel 'xlsx'"""

		self.hoja_nueva = self.wb.create_sheet(str(nombre_hoja_nueva))


		
		

	def escribir_titulo(self, titulos, fila, hoja_activa, col_ini = 0, ):
		"""Escribe los titulos en la fila y columna que se indique"""
		self.columna_inicial = col_ini
		self.wb.active =  hoja_activa

		if type(titulos) is list:
			for text_titulo in titulos:

				self.columna_inicial += 1
				titulo = self.wb.active.cell(row = fila , column = self.columna_inicial)
				titulo.value = (text_titulo)

		elif type(titulos) is str:
				titulo = self.wb.active.cell(row = fila, column =  self.columna_inicial)
				titulo.value = (titulos)
	
	
	def escribir_en_hoja(self, contenido_lista, col, hoja_activa, fila_ini=1):
		"""Escribe en la hoja que esta activa"""

		self.wb.active =  hoja_activa

		if type(contenido_lista) is list and len(contenido_lista) == 1:
			for conte in contenido_lista:
				
				for texto in conte:

					if type(texto) is list:
						for text in texto:
							fila_ini +=1
							texto_celda = self.wb.active.cell(row = fila_ini, column = col)
							texto_celda.value = (text)
					else:

						fila_ini +=1
						texto_celda = self.wb.active.cell(row = fila_ini, column = col)
						texto_celda.value = (texto)
		elif type(contenido_lista) is list and  len(contenido_lista) >= 2:
			self.escribir_varias_columnas(hoja_activa, contenido_lista, col, fila_ini)



	def escribir_varias_columnas(self, hoja_activa, listas, col, fila):
		self.wb.active =  hoja_activa
		
		col = 0
		for conte in listas:
			col +=1
			fila = 1
			
			if type(conte) is list:               
				for texto in conte:
					fila +=1                      

					#print(texto,"|fila|",fila, "|columna|",col)

					texto_celda = self.wb.active.cell(row = fila, column = col)
					texto_celda.value = (str(texto))        
				
				#print("leyendo Otra Lista")
			else:
				print("Escritas todas las columnas")

			  

				




	def guardar_archivo_log(self, ruta_guardar):

		#path_save = asksaveasfile(defaultextension = ".xlsx")
		self.wb.save(ruta_guardar.name)

		print("Guardado el log")


class ArchivoExcelAbrir:

	def __init__(self, documento):

		self.ruta = documento
		

		self.abrir_documento()
		self.obtener_hojas()


	def abrir_documento(self):
		self.documento_open =  load_workbook(self.ruta)
		print("Se Abrio el documento: ",
		 	  self.ruta.split("\\")[-1])




	def obtener_hojas(self):
		"""Obtiene los nombres de las hojas
			del archivo excel dado"""

		self.hojas_nombres = {}

		for hoja_nombre, hojas in  zip(self.documento_open.sheetnames,
									   self.documento_open.worksheets):
			self.hojas_nombres[hoja_nombre] = hojas

		self.hojas_lista = [h for h in self.documento_open.worksheets]

		return self.hojas_nombres




	def leer_titulos(self, hoja, linea):
		"""Obtiene titulos de columnas del
			archivo que se abre"""

		self.linea_i = linea
		self.linea_ccn = self.linea_i -1
		self.titulos_columnas = {}
		self.hoja = hoja 
		#LLama al metodo que obtiene los titulos cccn del IQ
		if self.ruta.split("\\")[-1].split("_")[0] == 'IQ':
			self.leer_titulos_ccn_iq(hoja, self.linea_ccn)


		if type(self.hoja) is str:

			hoja = self.hojas_nombres[self.hoja]
			titulos = hoja[self.linea_i]

			for titulo in range(len(titulos)):
				if titulos[titulo].value is not None:
					self.titulos_columnas[titulos[titulo].value] = titulos[titulo]



		elif type(self.hoja) is int:

			hoja = self.hojas_lista[self.hoja]
			titulos = hoja[self.linea_i]


			for titulo in range(len(titulos)):
				if titulos[titulo].value is not None:
					self.titulos_columnas[titulos[titulo].value] = titulos[titulo]

		return self.obtener_claves_celdas(self.hoja)



	def obtener_claves_celdas(self, hoja):

		self.claves_columnas     = {}
		self.columnas_i = {}

		if type(hoja) is str :
			columna_max = self.hojas_nombres[hoja].max_column
			fila_max 	 = self.hojas_nombres[hoja].max_row

			for (celda_text,
				 celda_nombres_clv) in self.titulos_columnas.items():

					celda          = str(celda_nombres_clv).split(".")[-1][:-2]
					self.claves_columnas[celda_text.strip(' ')]  = celda.strip('>')		#Clave de la columna 					
					self.columnas_i [celda_text.strip(' ')] = self.linea_i			#fila inicial de lectura

		elif type(hoja) == int and type(hoja) != float:
			columna_max = self.hojas_lista[hoja].max_column
			fila_max 	 = self.hojas_lista[hoja].max_row

			for (celda_text, celda_nombres_clv) in self.titulos_columnas.items():

					celda          = str(celda_nombres_clv).split(".")[-1][:-2]
					self.claves_columnas[celda_text.strip(' ')]  = celda.strip('>')		#Clave de la columna
					self.columnas_i [celda_text.strip(' ')] = self.linea_i			#Fila donde se empieza a leer


		return [self.claves_columnas, self.columnas_i]




	def leer_titulos_ccn_iq(self, hoja_lectura, linea):
		"""Obtiene los titulos del IQ NUMEROS DE CONCEPTOS
		   de Nómina"""
		self.titulos_ccn = {}

		

		if type(hoja_lectura) is str:
			hoja = self.hojas_nombres[hoja_lectura]
			titulos = hoja[linea]

			for titulo in range(len(titulos)):
				if titulos[titulo].value is not None:
					self.titulos_ccn[titulos[titulo].value] = titulos[titulo]



		elif type(hoja_lectura) is int:

			hoja = self.hojas_lista[hoja_lectura]
			titulos = hoja[linea]

			for titulo in range(len(titulos)):
				if titulos[titulo].value is not None:
					self.titulos_ccn[titulos[titulo].value] = titulos[titulo]


		self.obtener_claves_celdas_ccn(hoja_lectura)


	def obtener_claves_celdas_ccn(self, hoja):

		self.columnas_ccn     = {}
		self.columnas_i_ccn = {}



		if type(hoja) is str :
			columna_max = self.hojas_nombres[hoja].max_column
			fila_max 	 = self.hojas_nombres[hoja].max_row

			for (celda_text,
				 celda_nombres_clv) in self.titulos_ccn.items():

				celd = str(celda_text)

				celda= str(celda_nombres_clv).split(".")[-1][:-2]
				self.columnas_ccn[celd.strip(' ')]  = celda.strip('>')   #Claves de las columnas				
				self.columnas_i_ccn [celd.strip(' ')] = self.linea_i	#fila inicial de lectura

		elif type(hoja) == int and type(hoja) != float:
			columna_max = self.hojas_lista[hoja].max_column
			fila_max 	 = self.hojas_lista[hoja].max_row

			for (celda_text, celda_nombres_clv) in self.titulos_ccn.items():

				celd = str(celda_text)
				celda          = str(celda_nombres_clv).split(".")[-1][:-2]
				self.columnas_ccn[celd.strip(' ')]  = celda.strip('>')		#Claves de las columnas				
				self.columnas_i_ccn [celd.strip(' ')] = self.linea_i		#fila inicial de lectura