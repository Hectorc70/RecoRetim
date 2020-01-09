


from openpyxl import Workbook





class Log:
    """Crea Archivo log de las rutas de los archivos de 
       una carpeta """



    def __init__(self, titulo_hoja_default = "Hoja 1"):

        self.crear_archivo_xlsx()
        self.wb.active       = 0
        self.wb.active.title = str(titulo_hoja_default)


        
  



    def crear_archivo_xlsx(self):
        """Crea un archivo nuevo, excel 'xlsx'"""


        self.wb = Workbook()


    def crear_hoja_nueva(self, nombre_hoja_nueva):
        """Crea una hoja nueva en el archivo 
           excel 'xlsx'"""

        self.hoja_nueva = self.wb.create_sheet(str(nombre_hoja_nueva))


        
        

    def escribir_titulo(self, titulos, fila, hoja_activa, col_ini = 0, ):
        """Escribe los titulos en la fila y columna que se indique"""
        self.columna_inicial = col_ini
        self.wb.active =  hoja_activa

        if type(titulos) is list:
            for text_titulo in titulos:

                self.columna_inicial += 1
                titulo = self.wb.active.cell(row = fila , column = self.columna_inicial)
                titulo.value = (text_titulo)

        elif type(titulos) is str:
                titulo = self.wb.active.cell(row = fila, column =  self.columna_inicial)
                titulo.value = (titulos)
    
    
    def escribir_en_hoja(self, contenido_lista, col, hoja_activa, fila_ini=1):
        """Escribe en la hoja que esta activa"""

        self.wb.active =  hoja_activa

        if type(contenido_lista) is list and len(contenido_lista) == 1:
            for conte in contenido_lista:
                
                for texto in conte:

                    fila_ini +=1
                    texto_celda = self.wb.active.cell(row = fila_ini, column = col)
                    texto_celda.value = (texto)

        elif type(contenido_lista) is list and  len(contenido_lista) >= 2:
            self.escribir_varias_columnas(hoja_activa, contenido_lista, col, fila_ini)



    def escribir_varias_columnas(self, hoja_activa, listas, col, fila):

        self.wb.active = hoja_activa
        col = 0
        for conte in listas:
            col +=1
            fila = 1
            
            if type(conte) is list:               
                for texto in conte:
                    fila +=1                      

                    #print(texto,"|fila|",fila, "|columna|",col)

                    texto_celda = self.wb.active.cell(row = fila, column = col)
                    texto_celda.value = (texto)        
                
                #print("leyendo Otra Lista")
            else:
                print("Escritas todas las columnas")

              

                




    def guardar_archivo_log(self,ruta_guardar):

        #path_save = asksaveasfile(defaultextension = ".xlsx")
        self.wb.save(ruta_guardar.name)

        print("Guardado el log")







